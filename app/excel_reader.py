import openpyxl
import config


def get_inventory():
    """Read Inventory tab from DHM_Inventory_Database.xlsx (read-only).
    Returns list of dicts with customer-safe fields only."""
    try:
        wb = openpyxl.load_workbook(config.INVENTORY_DB, read_only=True, data_only=True)
    except FileNotFoundError:
        return []

    if 'Inventory' not in wb.sheetnames:
        wb.close()
        return []

    ws = wb['Inventory']
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return []

    headers = [str(h).strip() if h else '' for h in rows[0]]
    items = []
    for row in rows[1:]:
        rec = dict(zip(headers, row))
        # Only expose customer-safe fields
        items.append({
            'item_name': rec.get('Item Name', '') or '',
            'pack_size': rec.get('Pack Size', '') or '',
            'cases': rec.get('Cases', '') or '',
            'vendor': rec.get('Vendor', '') or '',
        })

    # Filter out empty rows
    items = [i for i in items if i['item_name']]
    return items
